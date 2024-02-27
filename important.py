import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
import time
import random
import pyperclip

names = []
city = []
name_button = None
city_button = None

def preprocess_numbers(numbers_text):
    processed_numbers = []
    for line in numbers_text.split('\n'):
        line = line.strip()
        if line:
            numbers = line.split()
            processed_numbers.extend(numbers)
    return processed_numbers

def read_excel(filename):
    global names, city
    try:
        wb = load_workbook(filename)
        sheet = wb.active
        numbers = []
        names = []
        city=[]
        found_phone_column = False
        found_name_column = False
        found_city_column = False
        
        for row in sheet.iter_rows(values_only=True):
            if not found_phone_column:
                if "Number" in row:
                    phone_index = row.index("Number")
                    found_phone_column = True
            if not found_name_column:
                if "Name" in row:
                    name_index = row.index("Name")
                    found_name_column = True
            if not found_city_column:
                if 'City' in row:
                    city_index = row.index("City")
                    found_city_column = True

        for row in sheet.iter_rows(values_only=True):
            if row[phone_index] is not None:
                numbers.append(row[phone_index] )
            if found_name_column:
                if row[name_index] is not None:
                    names.append(row[name_index])
                    if len(names) >=2:
                        add_name_button()
                else:
                    names.append('')
            if found_city_column:
                if row[city_index] is not None:
                    city.append(row[city_index])
                    if len(city) >= 2:
                        add_city_button()
                else:
                    city.append('')
        if "Number" in numbers:
            numbers.remove('Number')
        if "Name" in names:
            names.remove("Name")
        if 'City' in city:
            city.remove('City')

        numbers_entry.delete("1.0", "end")
        numbers_entry.insert("1.0", "\n".join(map(str, numbers)))
        return names,city
    except Exception as e:
        messagebox.showerror("Error", f"An error occurred while reading Excel file: {e}")

def add_name_button():
    global name_button
    name_button = ttk.Button(root, text="Add Name", command=insert_query_name)
    name_button.grid(row=1, column=2, padx=5, pady=5)

def add_city_button():
    global city_button
    city_button = ttk.Button(root, text="Add City", command=insert_query_city)
    city_button.grid(row=2, column=2, padx=0, pady=5)

def upload_excel():
    filename = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if filename:
        global names, city
        names, city = read_excel(filename)
        delete_button.config(state=tk.NORMAL)

def insert_query_name():
    cursor_pos = message_entry.index(tk.INSERT)
    message_entry.insert(cursor_pos, '{{query_name}}')

def insert_query_city():
    cursor_pos = message_entry.index(tk.INSERT)
    message_entry.insert(cursor_pos, '{{query_city}}')

def send_messages():
    global names, city
    numbers_text = numbers_entry.get("1.0", "end")
    original_message = message_entry.get("1.0", "end").strip()

    if not numbers_text or not original_message:
        messagebox.showerror("Error", "Please enter numbers and message")
        return
    numbers = preprocess_numbers(numbers_text)

    driver = None
    try:
        driver = webdriver.Chrome()
        driver.get('https://web.whatsapp.com')
        WebDriverWait(driver, 100).until(EC.presence_of_element_located((By.XPATH, "//div[@title='New chat']")))

        batch_size = 30
        pause_duration = 60

        for i in range(0, len(numbers), batch_size):
            batch_numbers = numbers[i:i+batch_size]
            if len(names) == 0 or len(city) == 0:
                for num in batch_numbers:
                    name = ''
                    city_name = ''
                    try:
                        send_message(driver, num, name, city_name, original_message)
                    except Exception as e:
                        print(f"Error sending message to {num}: {e}")
                        continue
            else:
                for num, name, city_name in zip(batch_numbers, names, city):
                    try:
                        send_message(driver, num, name, city_name, original_message)
                    except Exception as e:
                        print(f"Error sending message to {num}: {e}")
                        continue

            if i + batch_size < len(numbers):
                print(f"Sent messages to {i + batch_size} contacts. Taking a break for 1 minute...")
                time.sleep(pause_duration)
    except Exception as e:
        print("An error occurred:", e)
    finally:
        if driver:
            time.sleep(2)
            driver.quit()
            messagebox.showinfo("Success", "Messages sent successfully")

def send_message(driver, num, name, city_name, original_message):
    WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@title='New chat']")))
    new_chat_button = driver.find_element(By.XPATH, "//div[@title='New chat']")
    new_chat_button.click()

    search_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@contenteditable='true' and @role='textbox' and @title='Search input textbox']")))
    search_input.send_keys(num)
    time.sleep(0.5)

    loading = driver.find_elements(By.XPATH,"//span[@class='_11JPr' and contains(text(), 'Looking outside your contacts...')]")
    if loading:
        try:
            WebDriverWait(driver, 100).until(EC.invisibility_of_element_located((By.XPATH, "//span[@class='_11JPr' and contains(text(), 'Looking outside your contacts...')]")))
        except TimeoutException:
            print("Loading element did not disappear within the timeout period")

    in_contact = driver.find_elements(By.XPATH, "//div[@class='_2a-B5 VfC3c' and contains(text(), 'Contacts on WhatsApp')]")
    not_in_contact = driver.find_elements(By.XPATH, "//div[@class='_2a-B5 VfC3c' and contains(text(), 'Not in your contacts')]")
    not_in_whatsapp = driver.find_elements(By.XPATH, "//span[@class='_11JPr' and contains(text(), 'No results found')]")
    time.sleep(0.7)
    if in_contact:
        button = driver.find_element(By.XPATH, "//div[@tabindex='-1' and @role='button']")
        button.click()
    elif not_in_contact:
        button = driver.find_element(By.XPATH, "//div[@class='g0rxnol2 g0rxnol2 thghmljt p357zi0d rjo8vgbg ggj6brxn f8m0rgwh gfz4du6o ag5g9lrv bs7a17vp ov67bkzj']/div[@class='_2a-B5 VfC3c' and contains(text(), 'Not in your contacts')]/following-sibling::div[@tabindex='0' and @role='button']")
        button.click()
    elif not_in_whatsapp:
        cancel_button = driver.find_element(By.CSS_SELECTOR, "button[aria-label='Cancel search']")
        cancel_button.click()
        return
    time.sleep(0.5)
    message_input = WebDriverWait(driver, 10).until(EC.presence_of_element_located((By.XPATH, "//div[@class='to2l77zo gfz4du6o ag5g9lrv bze30y65 kao4egtt' and @contenteditable='true' and @role='textbox' and @title='Type a message']")))
    modified_message = original_message
    if name and '{{query_name}}' in modified_message:
        modified_message = modified_message.replace('{{query_name}}', name, 1)
    elif '{{query_name}}' in modified_message:
        modified_message = modified_message.replace('{{query_name}}', '', 1)

    if city_name and '{{query_city}}' in modified_message:
        modified_message = modified_message.replace('{{query_city}}', city_name, 1)
    elif '{{query_city}}' in modified_message:
        modified_message = modified_message.replace('{{query_city}}', '', 1)

    pyperclip.copy(modified_message)
    message_input.send_keys(Keys.CONTROL, 'v')
    message_input.send_keys(Keys.RETURN)
    waiting_time = random.randint(1, 5)
    time.sleep(waiting_time)

def clear_data():
    global names, city, name_button, city_button 
    names.clear()
    city.clear()
    numbers_entry.delete('1.0', tk.END)
    message_entry.delete('1.0', tk.END)
    if name_button:
        name_button.grid_remove()
        name_button = None
    if city_button:
        city_button.grid_remove()
        city_button = None
    delete_button.config(state=tk.DISABLED)

root = tk.Tk()
root.title("WhatsApp Message Sender")

# Styling
root.configure(bg="#ededed")

# Define styles
style = ttk.Style()

# Header
header_frame = ttk.Frame(root)
header_frame.grid(row=0, column=0, columnspan=3, padx=10, pady=10, sticky="ew")

header_label = ttk.Label(header_frame, text="WhatsApp Message Sender", font=("Helvetica", 16, "bold"), background='#ededed', foreground='#075e54')
header_label.pack()

# Main Content
content_frame = ttk.Frame(root)
content_frame.grid(row=1, column=0, columnspan=3, padx=10, pady=10)

# Numbers Section
numbers_label = ttk.Label(content_frame, text="Numbers:", font=("Helvetica", 12), background='#ededed')
numbers_label.grid(row=0, column=0, sticky="w", padx=10, pady=5)

numbers_entry = tk.Text(content_frame, width=40, height=4)
numbers_entry.grid(row=1, column=0, padx=10, pady=5)

upload_button = ttk.Button(content_frame, text="Upload Excel", command=upload_excel)
upload_button.grid(row=1, column=1, padx=10, pady=5, sticky="w")



delete_button = ttk.Button(content_frame, text="Delete", command=clear_data, state=tk.DISABLED, compound=tk.CENTER)
delete_button.grid(row=1, column=2, columnspan=1, pady=2)

# Message Section
message_label = ttk.Label(content_frame, text="Message:", font=("Helvetica", 12), background='#ededed')
message_label.grid(row=2, column=0, sticky="w", padx=10, pady=5)

message_entry = tk.Text(content_frame, width=40, height=10)
message_entry.grid(row=3, column=0, padx=10, pady=5)


# Send Button
send_button = ttk.Button(content_frame, text="Send Messages", command=send_messages)
send_button.grid(row=4, column=0, columnspan=2, pady=10)

root.mainloop()